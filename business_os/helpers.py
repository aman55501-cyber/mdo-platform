"""
Business OS — the three helpers borrowed from the trading console, re-created
here so that ``business.py`` and ``agents.py`` import cleanly with **no
trading-console dependency**.

  1. read_table   — Excel/CSV reader: bytes -> list of {header: value} rows
  2. normalise    — number normaliser: "₹1.5 Cr" / "12,34,567" / "8%" -> float
  3. notify / _alert / _fresh — ntfy + Telegram push, with de-dupe throttling

Hard rules (verbatim from the handoff), enforced in this module:
  - secrets live in .env only (read via os.getenv, never hard-coded)
  - never log a key (see _redact / no secret ever reaches a log line)
  - every threshold field passes the normaliser first (see business.py/agents.py)
"""

from __future__ import annotations

import csv
import io
import logging
import os
import re
import time
from typing import Any

log = logging.getLogger("business_os")


# ---------------------------------------------------------------------------
# 1. Excel / CSV reader
# ---------------------------------------------------------------------------

def read_table(data: bytes, filename: str = "") -> list[dict[str, Any]]:
    """Read an uploaded spreadsheet into a list of ``{header: value}`` rows.

    Supports ``.xlsx``/``.xlsm`` (via openpyxl) and ``.csv``/``.tsv`` (stdlib).
    The first non-empty row is treated as the header. Blank trailing columns
    and fully-empty rows are dropped. Values are returned as-is (str/number);
    numeric interpretation is the normaliser's job, not the reader's.
    """
    name = (filename or "").lower().strip()

    if name.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return _read_xlsx(data)
    if name.endswith((".csv", ".tsv", ".txt")):
        return _read_csv(data, tab=name.endswith(".tsv"))

    # No/unknown extension: sniff. Excel files are ZIP archives -> "PK".
    if data[:2] == b"PK":
        return _read_xlsx(data)
    return _read_csv(data)


def _read_xlsx(data: bytes) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise RuntimeError(
            "openpyxl is required to read .xlsx files (pip install openpyxl)"
        ) from exc

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    header = _first_header(rows_iter)
    if not header:
        wb.close()
        return []

    out: list[dict[str, Any]] = []
    for raw in rows_iter:
        row = _zip_row(header, raw)
        if row is not None:
            out.append(row)
    wb.close()
    return out


def _read_csv(data: bytes, tab: bool = False) -> list[dict[str, Any]]:
    text = _decode(data)
    if not tab:
        # Sniff the delimiter from the first kB; default to comma.
        sample = text[:1024]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","
    else:
        delimiter = "\t"

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    header = _first_header(reader)
    if not header:
        return []

    out: list[dict[str, Any]] = []
    for raw in reader:
        row = _zip_row(header, raw)
        if row is not None:
            out.append(row)
    return out


def _decode(data: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _first_header(rows_iter) -> list[str]:
    """Advance ``rows_iter`` to the first non-empty row and return it as
    trimmed, de-duplicated string headers."""
    for raw in rows_iter:
        cells = [_clean_cell(c) for c in (raw or [])]
        if any(c != "" for c in cells):
            return _dedupe_headers(cells)
    return []


def _dedupe_headers(cells: list[str]) -> list[str]:
    header: list[str] = []
    seen: dict[str, int] = {}
    for i, c in enumerate(cells):
        name = c or f"col_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        header.append(name)
    return header


def _zip_row(header: list[str], raw) -> dict[str, Any] | None:
    values = list(raw or [])
    if not any(_clean_cell(v) != "" for v in values):
        return None  # fully-empty row
    row: dict[str, Any] = {}
    for i, key in enumerate(header):
        val = values[i] if i < len(values) else None
        row[key] = _clean_cell(val, keep_type=True)
    return row


def _clean_cell(value: Any, keep_type: bool = False) -> Any:
    if value is None:
        return "" if not keep_type else ""
    if isinstance(value, str):
        return value.strip()
    if keep_type and isinstance(value, (int, float)):
        return value
    return str(value).strip()


# ---------------------------------------------------------------------------
# 2. Number normaliser
# ---------------------------------------------------------------------------

# Indian-scale multipliers. Order matters for the regex alternation below.
_SCALE = {
    "arab": 1_000_000_000,
    "cr": 10_000_000, "crore": 10_000_000, "crores": 10_000_000,
    "lakh": 100_000, "lakhs": 100_000, "lac": 100_000, "lacs": 100_000, "l": 100_000,
    "k": 1_000, "thousand": 1_000,
    "m": 1_000_000, "mn": 1_000_000, "million": 1_000_000,
    "bn": 1_000_000_000, "billion": 1_000_000_000,
}

# The num group must contain at least one digit, so stray separators (a lone
# "." left over from stripping "Rs.") are skipped rather than matched as zero.
_SCALE_RE = re.compile(
    r"(?P<num>-?(?:\d[\d,]*(?:\.\d+)?|\.\d+))\s*"
    r"(?P<unit>arab|crores?|cr|lakhs?|lacs?|l|thousand|k|"
    r"million|mn|m|billion|bn)?\b",
    re.IGNORECASE,
)


def normalise(value: Any, default: float | None = None) -> float | None:
    """Normalise a messy human/spreadsheet number into a plain float.

    Handles: ``₹1,23,456``, ``1.5 Cr``, ``50 lakh``, ``8%``, ``(1,200)`` for
    negatives, ``Rs. 2.5cr``, thousands separators (Indian or Western), and
    already-numeric inputs. Returns ``default`` (None by default) when nothing
    numeric can be extracted.

    **Every threshold field must pass through this before comparison** — that is
    a hard rule from the handoff.
    """
    if value is None:
        return default
    if isinstance(value, bool):
        return default
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    if s == "" or s in {"-", "--", "n/a", "na", "nil", "none", "—"}:
        return default

    # Accounting-style negatives: (1,200) -> -1200
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    is_percent = "%" in s

    # Strip currency symbols/words and stray characters, keep digits/.,-/scale.
    cleaned = re.sub(r"(?i)\b(rs\.?|inr|usd|eur|gbp)\b", " ", s)
    cleaned = cleaned.replace("₹", " ").replace("$", " ").replace("€", " ")
    cleaned = cleaned.replace("£", " ").replace("%", " ")

    m = _SCALE_RE.search(cleaned)
    if not m or not m.group("num"):
        return default

    num_str = m.group("num").replace(",", "")
    if num_str in {"", "-", ".", "-."}:
        return default
    try:
        num = float(num_str)
    except ValueError:
        return default

    unit = (m.group("unit") or "").lower()
    if unit in _SCALE:
        num *= _SCALE[unit]

    if negative:
        num = -num
    # Percent values are returned as the raw percentage number (8% -> 8.0),
    # callers decide whether to divide by 100.
    _ = is_percent
    return num


# ---------------------------------------------------------------------------
# 3. Notify / alert helpers  (ntfy + Telegram, with de-dupe throttling)
# ---------------------------------------------------------------------------

# In-memory throttle state: {alert_key: last_sent_epoch}
_ALERT_STATE: dict[str, float] = {}
_DEFAULT_TTL = 6 * 3600  # don't repeat the same alert within 6h


def _redact(s: str) -> str:
    """Best-effort redaction so a token can never reach a log line."""
    return re.sub(r"[A-Za-z0-9_\-]{16,}", "***", s or "")


def notify(message: str, title: str | None = None, priority: str | None = None,
           tags: str | None = None) -> bool:
    """Push a message to whatever channels are configured in .env.

    Channels (all optional, wired purely from env — secrets never hard-coded):
      - ntfy:     NTFY_TOPIC  (+ optional NTFY_URL, default https://ntfy.sh,
                  + optional NTFY_TOKEN for auth)
      - Telegram: TELEGRAM_BOT_TOKEN + TELEGRAM_CHAT_ID

    Returns True if at least one channel accepted the message. Never raises;
    never logs a secret.
    """
    sent = False
    sent |= _notify_ntfy(message, title, priority, tags)
    sent |= _notify_telegram(message, title)
    if not sent:
        log.info("notify: no channel configured; message dropped (len=%d)",
                 len(message or ""))
    return sent


def _notify_ntfy(message: str, title: str | None, priority: str | None,
                 tags: str | None) -> bool:
    topic = os.getenv("NTFY_TOPIC")
    if not topic:
        return False
    base = os.getenv("NTFY_URL", "https://ntfy.sh").rstrip("/")
    url = f"{base}/{topic}"
    headers: dict[str, str] = {}
    if title:
        headers["Title"] = title
    if priority:
        headers["Priority"] = priority
    if tags:
        headers["Tags"] = tags
    token = os.getenv("NTFY_TOKEN")
    if token:
        headers["Authorization"] = f"Bearer {token}"
    try:
        import httpx

        r = httpx.post(url, data=(message or "").encode("utf-8"),
                       headers=headers, timeout=10.0)
        r.raise_for_status()
        return True
    except Exception as exc:  # never let notification failure break the caller
        log.warning("ntfy notify failed: %s", _redact(str(exc)))
        return False


def _notify_telegram(message: str, title: str | None) -> bool:
    token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        return False
    text = f"*{title}*\n{message}" if title else (message or "")
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    try:
        import httpx

        r = httpx.post(
            url,
            json={"chat_id": chat_id, "text": text, "parse_mode": "Markdown"},
            timeout=10.0,
        )
        r.raise_for_status()
        return True
    except Exception as exc:
        # _redact keeps the bot token out of the log even if it appears in the URL
        log.warning("telegram notify failed: %s", _redact(str(exc)))
        return False


def _fresh(key: str, ttl: int = _DEFAULT_TTL) -> bool:
    """Return True if ``key`` has NOT fired within ``ttl`` seconds.

    A freshness gate: use before sending a repeatable alert so the same
    condition doesn't spam every scan cycle. Reading is side-effect free —
    call ``_alert`` to actually send and stamp the key.
    """
    last = _ALERT_STATE.get(key)
    if last is None:
        return True
    return (time.time() - last) >= ttl


def _alert(key: str, message: str, ttl: int = _DEFAULT_TTL,
           title: str | None = None, priority: str | None = None,
           tags: str | None = None) -> bool:
    """Send an alert at most once per ``ttl`` window per ``key``.

    Combines ``_fresh`` (throttle check) with ``notify`` (delivery) and stamps
    the key on a successful send. Returns True if the alert was sent this call.
    """
    if not _fresh(key, ttl):
        return False
    ok = notify(message, title=title, priority=priority, tags=tags)
    if ok:
        _ALERT_STATE[key] = time.time()
    return ok


def reset_alert_state() -> None:
    """Clear throttle state (used by tests / a manual 'resend everything')."""
    _ALERT_STATE.clear()
